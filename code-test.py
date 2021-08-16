import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb=load_workbook("Orders-With Nulls.xlsx")
wb1=Workbook()
ws = wb.active
ws1 = wb1.active
#sheets=wb.sheetnames
#print(wb.active.title)#to know sheet name.
sh1=wb['Orders']
ws1.title="Orders"
row=sh1.max_row
column=sh1.max_column
#print(row,column) #to check row and column count.

for i in range(1,row+1):
    for j in range(2,5,2):
        char = get_column_letter(j)
        print(ws[char + str(i)].value)
wb1.save("Result_file.xlsx")
